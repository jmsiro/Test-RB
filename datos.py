from openpyxl import Workbook, load_workbook
from itertools import count #Count es una función con 2 argumentos (start, step), que crea un itinerador - Use esta en lugar de un range ya que la función para obtener la ultima fila y ultima columna no devolvian el resultado correcto.

def armado_matriz(wb):
    # Abro libro. Los flags indican que es modo lectura y que muesta los resultados de las formulas del excel.
    libro = load_workbook(filename=wb, read_only=False, data_only=True)

    #En este caso que es una sola hoja tomo la hoja activa, pero en caso que haya mas de una deberia seleccionarla utilizando el nombre de la misma como si fuera un diccionario.
    hoja = libro['Hoja1']

    matriz = []
    datos = {}

    # Loop 1 = Recorre filas
    # Loop 2 = Recorre Columnas
    # Se crea una Lista de Diccionarios donde las 'Keys' son los titulos de las columnas y el 'Value' el valor de esa columna para cada fila.
    for r in count(start=2, step=1):
        if hoja.cell(row=r, column=1).value == None:
            break
        else:
            datos['Fila']=r
            for c in range(1,12):
                if hoja.cell(row=1, column=c).value == None:
                    matriz.append(datos.copy()) # Se genera una copia del diccionario y se incopora a la lista
                    datos.clear() # Se limpia del diccionario original para la proxima vuelta del loop
                    break
                else:
                    datos[hoja.cell(row=1, column=c).value] = hoja.cell(row=r, column=c).value
    
    return matriz